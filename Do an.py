import cv2
import pytesseract
import openpyxl
wb = openpyxl.load_workbook("C:\\Users\\ACER\\Downloads\\File excel\\Sample.xlsx")

sheet = wb['Sheet1']
cell = sheet.cell(row=1, column=1,value = "LastNumber")
cell = sheet.cell(row=1, column=2,value = "FirstNumber")
cell = sheet.cell(row=1, column=3,value = "Score")

pytesseract.pytesseract.tesseract_cmd = 'C:\\Program Files\\Tesseract-OCR\\tesseract.exe'

def camera1():
    LastNumber = None
    ListNumber1 = []

    video1 = cv2.VideoCapture("C:\\Users\\ACER\\OneDrive\\Máy tính\\TestCam.webm")
    video1.set(3, 640)
    video1.set(4, 480)

# Allows continuous frames
    while True:
    # Capture each frame from the video feed
        _, frames = video1.read()
        try:
            data4 = pytesseract.image_to_data(frames)
        except:
            continue
        for z, a in enumerate(data4.splitlines()):
        # Counter
            if z != 0:
            # Converts 'data1' string into a list stored in 'a'
            # print(len(a))
                a = a.split()
            # Checking if array contains a word
                if len(a) == 12:
                # Storing values
                    x, y = int(a[6]), int(a[7])
                    w, h = int(a[8]), int(a[9])
                    # Display bounding box of each word
                    cv2.rectangle(frames, (x, y), (x + w, y + h), (0, 0, 255), 2)
                    # Display detected word under each bounding box
                    cv2.putText(frames, a[11], (x - 15, y),
                                cv2.FONT_HERSHEY_PLAIN, 2, (255, 0, 0), 1)
                    grayImg = cv2.cvtColor(frames, cv2.COLOR_BGR2GRAY)
                    if a[11].isdigit():
                        if (a[11] != LastNumber):
                            LastNumber = a[11]
                            ListNumber1.append(LastNumber)
                            path = "C:\\Users\\ACER\\Downloads\\Video Capture\\" + str(a[11]) +".jpg"
                            cv2.imwrite(path, grayImg)
                        break
        cv2.imshow("Camera 1",frames)
        if cv2.waitKey(1) & 0xFF == ord('q'):
            video1.release()
            cv2.destroyAllWindows()
            break
    return ListNumber1

def camera2():
    video2 = cv2.VideoCapture("C:\\Users\\ACER\\OneDrive\\Máy tính\\TestCam.webm")
    FirstNumber = None
    ListNumber2 = []
    video2.set(3, 640)
    video2.set(4, 480)

    # Allows continuous frames
    while True:

        # Capture each frame from the video feed
        _, frame = video2.read()
        try:
            data5 = pytesseract.image_to_data(frame)
        except:
            continue
        for number, b in enumerate(data5.splitlines()):
            # Counter
            if number != 0:
                # Converts 'data1' string into a list stored in 'a'
                # print(len(a))
                b = b.split()
                # Checking if array contains a word
                if len(b) == 12:
                    # Storing values 
                    xid, yid = int(b[6]), int(b[7])
                    wid, hid = int(b[8]), int(b[9])
                    # Display bounding box of each word
                    cv2.rectangle(frame, (xid, yid), (xid + wid, yid + hid), (0, 0, 255), 2)
                    # Display detected word under each bounding box
                    cv2.putText(frame, b[11], (xid - 15, yid),
                                cv2.FONT_HERSHEY_PLAIN, 2, (255, 0, 0), 1)
                    grayImg = cv2.cvtColor(frame, cv2.COLOR_BGR2GRAY)
                    if b[11].isdigit():
                        if (b[11] != FirstNumber):
                            FirstNumber = b[11]
                            ListNumber2.append(FirstNumber)
                            path = "C:\\Users\\ACER\\Downloads\\Video Capture2\\" + str(b[11]) +".jpg"
                            cv2.imwrite(path, grayImg)
                        break
        cv2.imshow("Camera 2",frame)
        if cv2.waitKey(1) & 0xFF == ord('q'):
            video2.release()
            cv2.destroyAllWindows()
            break
    return ListNumber2
def camera3():
    video3 = cv2.VideoCapture("C:\\Users\\ACER\\OneDrive\\Máy tính\\TestCam.webm")
    video3.set(3, 640)
    video3.set(4, 480)
    while True:

        ret , framess = video3.read()
        if ret == False:
            break
        cv2.imshow("Camera 3",framess)
        if cv2.waitKey(1) & 0xFF == ord('q'):
            video3.release()
            cv2.destroyAllWindows()
            break
    return 
camera01= camera1()

camera02= camera2()

start_row =0
for i in range(len((camera01))):
    for j in range(len(camera02)):
        if i == j:
            if camera01[i] == camera02[j]:
                start_row += 1
                sheet.cell(row = start_row +1 , column=1,value =camera01[i]) 
                sheet.cell(row = start_row +1 , column=2,value =camera02[j])
            else:
                print("Số phách không trùng" )
        else:
            continue
wb.save("C:\\Users\\ACER\\Downloads\\File excel\\Sample.xlsx")






